"""
Genera el reporte de ventas diario de Printy (Cuadros / Foto Libros / Copias / Otros).

Uso:
    python generar_reporte.py [--dias 30] [--desde 2026-07-01] [--hasta 2026-08-18]

Por defecto genera el reporte de los ultimos 30 dias.
Requiere: pandas, openpyxl  (pip install pandas openpyxl)

Credenciales WooCommerce: por defecto leidas de data/datos.txt (mismo formato
de siempre: linea 2 = consumer key ck_..., linea 5 = consumer secret cs_...).
Si existen las variables de entorno WC_CONSUMER_KEY / WC_CONSUMER_SECRET,
tienen prioridad sobre el archivo (para correr sin ese archivo, p.ej. en un
entorno en la nube).
"""
import argparse
import json
import os
import re
import time
import urllib.request
from datetime import datetime, timedelta
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

ROOT = Path(__file__).resolve().parent.parent
DATOS_TXT = ROOT / 'data' / 'datos.txt'
BASE = 'https://printy.photos/wp-json/wc/v3'

CUADROS_IDS = {36, 38}     # Cuadros + Cuadros en Tela
FOTOLIBROS_IDS = {37}
COPIAS_IDS = {35}
BUCKETS_MAIN = ['Cuadros', 'Foto Libros', 'Copias']
VALID_STATUSES = {'processing', 'completed', 'on-hold'}

HEADER_FILL = PatternFill('solid', fgColor='1F2937')
HEADER_FONT = Font(color='FFFFFF', bold=True)
TOTAL_FILL = PatternFill('solid', fgColor='E5E7EB')
TOTAL_FONT = Font(bold=True)
THIN = Side(style='thin', color='D1D5DB')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CURRENCY_FMT = '#,##0'


def load_credentials():
    env_ck = os.environ.get('WC_CONSUMER_KEY')
    env_cs = os.environ.get('WC_CONSUMER_SECRET')
    if env_ck and env_cs:
        return env_ck, env_cs
    text = DATOS_TXT.read_text(encoding='utf-8')
    ck = re.search(r'ck_[a-f0-9]+', text).group(0)
    cs = re.search(r'cs_[a-f0-9]+', text).group(0)
    return ck, cs


def get(ck, cs, path, params, retries=4):
    qs = '&'.join(f'{k}={v}' for k, v in params.items())
    url = f'{BASE}/{path}?{qs}&consumer_key={ck}&consumer_secret={cs}'
    for attempt in range(retries):
        try:
            with urllib.request.urlopen(urllib.request.Request(url)) as r:
                # r.headers (email.message.Message) does case-insensitive
                # lookups; converting to a plain dict loses that and silently
                # breaks X-WP-TotalPages lookups when a proxy/CDN lowercases
                # header names (observed from some cloud egress paths, not
                # from a home network) - keep the Message object instead of
                # dict(r.headers).
                return json.loads(r.read().decode('utf-8')), r.headers
        except urllib.error.HTTPError as e:
            if e.code in (429, 502, 503, 504) and attempt < retries - 1:
                time.sleep(2 ** attempt)
                continue
            raise


def fetch_all(ck, cs, path, params):
    page = 1
    results = []
    while True:
        p = dict(params, page=page, per_page=100)
        d, headers = get(ck, cs, path, p)
        if not d:
            break
        results.extend(d)
        total_pages = int(headers.get('X-WP-TotalPages', '1'))
        if page >= total_pages:
            break
        page += 1
        time.sleep(0.3)  # avoid tripping the host's rate limiting on long paginated pulls
    return results


def fetch_data(ck, cs, desde, hasta):
    products = fetch_all(ck, cs, 'products', {'status': 'any'})
    orders = fetch_all(ck, cs, 'orders', {
        'after': f'{desde}T00:00:00',
        'before': f'{hasta}T23:59:59',
        'status': 'any',
    })
    return products, orders


def bucket_for(prod_cat, prod_catnames, product_id):
    cats = prod_cat.get(product_id, set())
    if cats & CUADROS_IDS:
        return 'Cuadros'
    if cats & FOTOLIBROS_IDS:
        return 'Foto Libros'
    if cats & COPIAS_IDS:
        return 'Copias'
    if cats:
        return 'Otros (' + prod_catnames.get(product_id, ['?'])[0] + ')'
    return 'Otros (sin categoria)'


def build_dataframe(products, orders):
    prod_cat = {p['id']: {c['id'] for c in p.get('categories', [])} for p in products}
    prod_catnames = {p['id']: [c['name'] for c in p.get('categories', [])] for p in products}

    rows = []
    for o in orders:
        if o['status'] not in VALID_STATUSES:
            continue
        date_str = o['date_created'][:10]
        for li in o['line_items']:
            pid = li['product_id']
            bucket = bucket_for(prod_cat, prod_catnames, pid)
            meta_all = {m.get('key'): m.get('value') for m in li.get('meta_data', [])}
            meta = {k: v for k, v in meta_all.items()
                     if k and not str(k).startswith('wc_imaxel') and k != 'proyecto'}
            tamano = meta.get('pa_tamano') or ''
            cobertura = meta.get('pa_cobertura') or ''
            base_name = li['name'].split(' - ')[0]
            tipo_parts = [base_name]
            if tamano:
                tipo_parts.append(str(tamano))
            if cobertura:
                tipo_parts.append(str(cobertura))

            # Real photo-copy count, Copias category only (the same meta
            # keys appear on other categories' line items for unrelated
            # reasons — e.g. a handful of Foto Libros lines also carry the
            # 'pages_copies' key — so this is deliberately scoped to bucket
            # == 'Copias'). Two eras coexist in the order history:
            # - Current flow (Copias por unidad / Set de Copias): cart qty is
            #   always 1 per design project; the real print count lives in
            #   this imaxel meta field.
            # - Older flow (pre photo-editor-redesign, tagged with the
            #   'imaxel_editors_jobs_ids' meta): the customer set the cart
            #   quantity directly to the number of prints wanted, so qty IS
            #   the real count there (verified against per-unit pricing).
            # Neither meta present (current Estilo Vintage, or photo albums)
            # -> unknown, don't guess.
            copias = None
            if bucket == 'Copias':
                copias_raw = meta_all.get('wc_imaxel_project_design_pages_copies')
                if copias_raw not in (None, ''):
                    try:
                        copias = int(copias_raw)
                    except (TypeError, ValueError):
                        copias = None
                elif meta_all.get('imaxel_editors_jobs_ids') not in (None, ''):
                    copias = li['quantity']

            rows.append({
                'fecha': date_str,
                'bucket': bucket,
                'producto': base_name,
                'tipo': ' - '.join(tipo_parts),
                'cantidad': li['quantity'],
                'valor': float(li['total']),
                'copias': copias,
                'order_id': o['id'],
            })
    df = pd.DataFrame(rows)
    if len(df):
        df['fecha'] = pd.to_datetime(df['fecha'])
    return df


def style_header_row(ws, row, ncols, start_col=1):
    for c in range(start_col, start_col + ncols):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = BORDER


def style_total_row(ws, row, ncols, start_col=1):
    for c in range(start_col, start_col + ncols):
        cell = ws.cell(row=row, column=c)
        cell.fill = TOTAL_FILL
        cell.font = TOTAL_FONT
        cell.border = BORDER


def autofit(ws, min_width=10, max_width=42):
    widths = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            widths[cell.column_letter] = max(widths.get(cell.column_letter, 0), len(str(cell.value)))
    for col, w in widths.items():
        ws.column_dimensions[col].width = max(min_width, min(max_width, w + 2))


def write_df(ws, start_row, start_col, data, index_label=None, currency_cols=None):
    currency_cols = currency_cols or set()
    nrows, ncols_data = data.shape
    ncols = ncols_data + 1
    ws.cell(row=start_row, column=start_col, value=index_label or '')
    for j, col in enumerate(data.columns):
        ws.cell(row=start_row, column=start_col + 1 + j, value=str(col))
    style_header_row(ws, start_row, ncols, start_col)
    for i, (idx, r) in enumerate(data.iterrows()):
        rr = start_row + 1 + i
        idx_val = idx.strftime('%Y-%m-%d') if hasattr(idx, 'strftime') else idx
        cidx = ws.cell(row=rr, column=start_col, value=idx_val)
        cidx.border = BORDER
        cidx.font = Font(bold=True)
        for j, col in enumerate(data.columns):
            val = r[col]
            cc = ws.cell(row=rr, column=start_col + 1 + j)
            cc.border = BORDER
            if pd.isna(val) or val == 0:
                cc.value = None
            else:
                cc.value = float(val) if isinstance(val, (int, float)) else val
                if col in currency_cols or currency_cols == 'ALL':
                    cc.number_format = CURRENCY_FMT
    return start_row + 1 + nrows


def daily_summary(sub):
    return sub.groupby('fecha').agg(cantidad=('cantidad', 'sum'), valor=('valor', 'sum')).sort_index(ascending=False)


def pivot_tipo(sub, value_col):
    order = sub.groupby('tipo')['valor'].sum().sort_values(ascending=False).index.tolist()
    p = sub.pivot_table(index='fecha', columns='tipo', values=value_col, aggfunc='sum', fill_value=0)
    return p.reindex(columns=order).sort_index(ascending=False)


def build_category_sheet(wb, name, sub):
    has_copias = 'copias' in sub.columns and sub['copias'].notna().any()

    ws = wb.create_sheet(name)
    ws['A1'] = f'{name} — detalle diario'
    ws['A1'].font = Font(bold=True, size=14)
    subtitle = f"{int(sub['cantidad'].sum())} unidades  |  ${sub['valor'].sum():,.0f} en el periodo"
    if has_copias:
        subtitle += f"  |  {int(sub['copias'].sum())} copias reales (Set de Copias + Copias por unidad)"
    ws['A2'] = subtitle
    ws['A2'].font = Font(italic=True, size=9, color='555555')

    row = 4
    ws.cell(row=row, column=1, value='Resumen diario (todos los tipos)').font = Font(bold=True, size=11)
    row += 1
    ds = daily_summary(sub)
    ds.columns = ['Cantidad', 'Valor']
    row = write_df(ws, row, 1, ds, index_label='Fecha', currency_cols={'Valor'}) + 2

    ws.cell(row=row, column=1, value='Cantidad de pedidos por tipo y dia').font = Font(bold=True, size=11)
    row += 1
    row = write_df(ws, row, 1, pivot_tipo(sub, 'cantidad'), index_label='Fecha') + 2

    ws.cell(row=row, column=1, value='Valor ($) por tipo y dia').font = Font(bold=True, size=11)
    row += 1
    row = write_df(ws, row, 1, pivot_tipo(sub, 'valor'), index_label='Fecha', currency_cols='ALL') + 2

    if has_copias:
        ws.cell(row=row, column=1, value='Copias reales por tipo y dia').font = Font(bold=True, size=11)
        row += 1
        ws.cell(row=row, column=1, value='Solo Set de Copias y Copias por unidad tienen este dato; Estilo Vintage y Álbumes no lo registran (ver nota abajo).').font = Font(italic=True, size=8, color='888888')
        row += 1
        copias_sub = sub[sub['copias'].notna()]
        if len(copias_sub):
            write_df(ws, row, 1, pivot_tipo(copias_sub, 'copias'), index_label='Fecha')

    ws.freeze_panes = 'B5'
    autofit(ws)


def build_workbook(df, desde, hasta):
    wb = Workbook()
    wb.remove(wb.active)

    ws = wb.create_sheet('Resumen')
    ws['A1'] = 'Reporte de ventas diario — Printy'
    ws['A1'].font = Font(bold=True, size=14)
    ws['A2'] = (f"Periodo: {desde} a {hasta}  |  Estados incluidos: processing, completed, "
                "on-hold (excluye cancelados/fallidos)")
    ws['A2'].font = Font(italic=True, size=9, color='555555')

    bucket_order = df.groupby('bucket')['valor'].sum().sort_values(ascending=False).index.tolist()
    bdq = df.pivot_table(index='fecha', columns='bucket', values='cantidad', aggfunc='sum', fill_value=0)
    bdq = bdq.reindex(columns=bucket_order).sort_index(ascending=False)
    bdv = df.pivot_table(index='fecha', columns='bucket', values='valor', aggfunc='sum', fill_value=0)
    bdv = bdv.reindex(columns=bucket_order).sort_index(ascending=False)

    row = 4
    ws.cell(row=row, column=1, value='CANTIDAD por categoria y dia').font = Font(bold=True, size=11)
    row += 1
    row = write_df(ws, row, 1, bdq, index_label='Fecha') + 2
    ws.cell(row=row, column=1, value='VALOR ($) por categoria y dia').font = Font(bold=True, size=11)
    row += 1
    row = write_df(ws, row, 1, bdv, index_label='Fecha', currency_cols='ALL') + 2

    totals = df.groupby('bucket').agg(cantidad=('cantidad', 'sum'), valor=('valor', 'sum')).sort_values('valor', ascending=False)
    ws.cell(row=row, column=1, value='TOTAL del periodo por categoria').font = Font(bold=True, size=11)
    row += 1
    for j, h in enumerate(['Categoria', 'Cantidad', 'Valor']):
        ws.cell(row=row, column=j + 1, value=h)
    style_header_row(ws, row, 3)
    row += 1
    grand_qty = grand_val = 0
    for cat, r in totals.iterrows():
        ws.cell(row=row, column=1, value=cat).border = BORDER
        ws.cell(row=row, column=2, value=int(r['cantidad'])).border = BORDER
        c = ws.cell(row=row, column=3, value=float(r['valor']))
        c.number_format = CURRENCY_FMT
        c.border = BORDER
        grand_qty += r['cantidad']
        grand_val += r['valor']
        row += 1
    ws.cell(row=row, column=1, value='TOTAL GENERAL')
    ws.cell(row=row, column=2, value=int(grand_qty))
    c = ws.cell(row=row, column=3, value=float(grand_val))
    c.number_format = CURRENCY_FMT
    style_total_row(ws, row, 3)
    ws.freeze_panes = 'B5'
    autofit(ws)

    for b in BUCKETS_MAIN:
        sub = df[df['bucket'] == b]
        if len(sub):
            build_category_sheet(wb, b, sub)

    otros = df[~df['bucket'].isin(BUCKETS_MAIN)].copy()
    if len(otros):
        otros['tipo'] = otros['bucket'].str.replace(r'Otros \(', '', regex=True).str.replace(r'\)', '', regex=True) + ' - ' + otros['producto']
        build_category_sheet(wb, 'Otros', otros)

    ws = wb.create_sheet('Detalle')
    headers = ['Fecha', 'Categoria', 'Producto', 'Tipo (detalle)', 'Cantidad', 'Valor', '# Pedido']
    for j, h in enumerate(headers):
        ws.cell(row=1, column=j + 1, value=h)
    style_header_row(ws, 1, len(headers))
    d2 = df.sort_values('fecha', ascending=False)
    for i, r in enumerate(d2.itertuples(index=False)):
        rr = i + 2
        ws.cell(row=rr, column=1, value=r.fecha.strftime('%Y-%m-%d'))
        ws.cell(row=rr, column=2, value=r.bucket)
        ws.cell(row=rr, column=3, value=r.producto)
        ws.cell(row=rr, column=4, value=r.tipo)
        ws.cell(row=rr, column=5, value=int(r.cantidad))
        c = ws.cell(row=rr, column=6, value=float(r.valor))
        c.number_format = CURRENCY_FMT
        ws.cell(row=rr, column=7, value=int(r.order_id))
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f"A1:G{len(d2) + 1}"
    autofit(ws)

    wb.move_sheet('Resumen', offset=-len(wb.sheetnames))
    return wb


# ============================================================
# HTML dashboard (Resumen / Cuadros / Foto Libros / Copias / Otros / Detalle)
# ============================================================

def build_report_json(df, desde, hasta):
    """Aggregate the line-item dataframe into the JSON structure the HTML
    template (report_template.html) expects: daily totals per category,
    per-category daily rows + a type breakdown, a type ranking, and the raw
    detail rows."""
    rows = df.copy()
    rows['fecha'] = rows['fecha'].dt.strftime('%Y-%m-%d')
    rows['bucket_display'] = rows['bucket']
    is_other = ~rows['bucket'].isin(BUCKETS_MAIN)
    rows.loc[is_other, 'bucket'] = 'Otros'

    d0 = datetime.strptime(desde, '%Y-%m-%d')
    d1 = datetime.strptime(hasta, '%Y-%m-%d')
    all_dates = []
    d = d0
    while d <= d1:
        all_dates.append(d.strftime('%Y-%m-%d'))
        d += timedelta(days=1)

    cat_keys = BUCKETS_MAIN + ['Otros']
    daily = {dt: {k: {'qty': 0, 'valor': 0.0} for k in cat_keys} for dt in all_dates}
    orders_by_day = {dt: set() for dt in all_dates}
    for r in rows.itertuples(index=False):
        if r.fecha not in daily:
            continue
        daily[r.fecha][r.bucket]['qty'] += r.cantidad
        daily[r.fecha][r.bucket]['valor'] += r.valor
        orders_by_day[r.fecha].add(r.order_id)
    # pedidos = distinct orders that day, across all categories (an order
    # with lines in two categories counts once, not twice) - each order has
    # a single date_created, so summing this across a date range never
    # double-counts an order the way summing per-category totals would.
    daily_series = [{'fecha': dt, 'pedidos': len(orders_by_day[dt]), **daily[dt]} for dt in all_dates]

    categories_out = []
    for key in cat_keys:
        sub = rows[rows['bucket'] == key]
        total_qty = int(sub['cantidad'].sum())
        total_valor = float(sub['valor'].sum())
        days_con_venta = sub['fecha'].nunique()

        tracks_copias = bool(sub['copias'].notna().any())

        daily_rows = []
        for dt, grp in sub.groupby('fecha'):
            type_agg = grp.groupby('tipo').agg(qty=('cantidad', 'sum'), valor=('valor', 'sum'), copias=('copias', 'sum'))
            types_sorted = [
                {'tipo': t, 'qty': int(v['qty']), 'valor': float(v['valor']), 'copias': int(v['copias'])}
                for t, v in type_agg.sort_values('valor', ascending=False).iterrows()
            ]
            daily_rows.append({
                'fecha': dt,
                'qty': int(grp['cantidad'].sum()),
                'valor': float(grp['valor'].sum()),
                'copias': int(grp['copias'].sum()),
                'types': types_sorted,
            })
        daily_rows.sort(key=lambda r: r['fecha'], reverse=True)

        rank_agg = sub.groupby('tipo').agg(qty=('cantidad', 'sum'), valor=('valor', 'sum'), copias=('copias', 'sum')).sort_values('valor', ascending=False)
        ranking = []
        for t, v in rank_agg.iterrows():
            pct = round(100 * v['valor'] / total_valor, 1) if total_valor else 0
            ranking.append({'tipo': t, 'qty': int(v['qty']), 'valor': float(v['valor']), 'copias': int(v['copias']), 'pct': pct})

        categories_out.append({
            'key': key,
            'label': key,
            'total_qty': total_qty,
            'total_valor': total_valor,
            'total_copias': int(sub['copias'].sum()),
            'tracks_copias': tracks_copias,
            'days_con_venta': int(days_con_venta),
            'avg_ticket': round(total_valor / total_qty, 0) if total_qty else 0,
            'daily_rows': daily_rows,
            'ranking': ranking,
        })

    detalle = rows.sort_values(['fecha', 'order_id'], ascending=False)
    detalle_out = [
        {'fecha': r.fecha, 'categoria': r.bucket, 'categoria_detalle': r.bucket_display,
         'producto': r.producto, 'tipo': r.tipo, 'cantidad': int(r.cantidad),
         'valor': float(r.valor), 'copias': (None if pd.isna(r.copias) else int(r.copias)),
         'pedido': int(r.order_id)}
        for r in detalle.itertuples(index=False)
    ]

    return {
        'desde': desde,
        'hasta': hasta,
        'generated_at': datetime.now().strftime('%Y-%m-%d %H:%M'),
        'categories': categories_out,
        'daily_series': daily_series,
        'detalle': detalle_out,
    }


def build_html(df, desde, hasta, template_path=None):
    template_path = Path(template_path) if template_path else Path(__file__).resolve().parent / 'report_template.html'
    template = template_path.read_text(encoding='utf-8')
    report_data = build_report_json(df, desde, hasta)
    data_json = json.dumps(report_data, ensure_ascii=False).replace('</script', '<\\/script')
    return template.replace('__REPORT_DATA_JSON__', data_json)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--dias', type=int, default=30, help='Cantidad de dias hacia atras (ignorado si se pasa --desde)')
    ap.add_argument('--desde', type=str, default=None, help='Fecha inicio YYYY-MM-DD')
    ap.add_argument('--hasta', type=str, default=None, help='Fecha fin YYYY-MM-DD (default: hoy)')
    ap.add_argument('--formato', choices=['xlsx', 'html', 'both'], default='xlsx', help='Formato de salida (default: xlsx)')
    ap.add_argument('--out', type=str, default=None, help='Ruta del archivo de salida (si se genera un solo formato)')
    args = ap.parse_args()

    hasta = args.hasta or datetime.now().strftime('%Y-%m-%d')
    if args.desde:
        desde = args.desde
    else:
        desde = (datetime.strptime(hasta, '%Y-%m-%d') - timedelta(days=args.dias)).strftime('%Y-%m-%d')

    ck, cs = load_credentials()
    print(f'Descargando datos de WooCommerce ({desde} a {hasta})...')
    products, orders = fetch_data(ck, cs, desde, hasta)
    print(f'{len(products)} productos, {len(orders)} pedidos descargados')

    df = build_dataframe(products, orders)
    print(f'{len(df)} lineas de pedido validas (processing/completed/on-hold)')

    if args.formato in ('xlsx', 'both'):
        wb = build_workbook(df, desde, hasta)
        out_path = Path(args.out) if (args.out and args.formato == 'xlsx') else ROOT / 'reportes' / f'Reporte_Ventas_{desde}_a_{hasta}.xlsx'
        out_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(out_path)
        print(f'Guardado: {out_path}')

    if args.formato in ('html', 'both'):
        html = build_html(df, desde, hasta)
        out_path = Path(args.out) if (args.out and args.formato == 'html') else ROOT / 'reportes' / 'Reporte_Ventas_Printy.html'
        out_path.parent.mkdir(parents=True, exist_ok=True)
        out_path.write_text(html, encoding='utf-8')
        print(f'Guardado: {out_path}')


if __name__ == '__main__':
    main()
